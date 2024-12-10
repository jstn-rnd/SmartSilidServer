from datetime import timezone
import subprocess
import time

from django.shortcuts import render, redirect

from django.views.decorators.csrf import csrf_exempt

from django.views.decorators.http import require_POST

from django.shortcuts import render, HttpResponse, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from wakeonlan import send_magic_packet

from server_app.views_wol import get_ip_from_mac, normalize_mac
from .models import Computer, Semester, Schedule, RfidLogs, ClassInstance, Attendance, User, RFID, User, Scan, Student
from .forms import ScheduleForm

from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from datetime import datetime as datetime2
from django.utils.dateparse import parse_date
from datetime import date
import math
import datetime
from datetime import timedelta
from .utils import format_fullname, format_fullname_lastname_first

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def view_records(request):
    
    rfid = RFID.objects.all()
    
    context = {
        'tests': rfid,
    }
    
    return render(request, 'server_app/view_records.html', context)


@require_POST
def update_approve_status(request):
    test_id = request.POST.get('test_id')
    if test_id:
        try:
            test = Test.objects.get(id=test_id)
            approve_value = int(request.POST.get(f'approve_{test_id}', 0))
            test.approved = approve_value  # Manually approve or deny
            test.save()
        except Test.DoesNotExist:
            pass
    return redirect('view_records')



from django.shortcuts import render, redirect
from django.utils import timezone
from .models import Schedule
from .forms import ScheduleForm
import pytz

def manage_schedules(request):
    if request.method == 'POST':
        
        print("POST request received")
        schedule_form = ScheduleForm(request.POST)
        if schedule_form.is_valid():
            print("Form is valid")
            schedule = schedule_form.save()
            print(f"Schedule saved: {schedule}")
            update_rfids_approval(schedule)  # Call the function to update RFIDs approval
            print("RFID approval update called")
            return redirect('manage_schedules')
        else:
            print("Form is not valid")
    else:
        schedule_form = ScheduleForm()

    schedules = Schedule.objects.all()
    rfids = RFID.objects.values_list('RFID', flat=True)
    return render(request, 'server_app/manage_schedules.html', {
        'schedule_form': schedule_form,
        'schedules': schedules,
        'rfids': rfids,
    })

def update_rfids_approval(schedule, rfid):
    local_tz = pytz.timezone('Asia/Manila')
    local_time = timezone.now().astimezone(local_tz)
    current_time = local_time.time()
    current_weekday = local_time.strftime('%a')

    faculty = schedule.faculty
    if schedule.start_time <= current_time <= schedule.end_time and current_weekday in schedule.get_weekdays_display():

        for rfid_instance in faculty.rfids.all(): 
            
            if rfid_instance.rfid == rfid:  
                rfid_instance.approved = 1
                rfid_instance.save()
                return schedule.id
            
            else : 
                print("Not match")
            
    else:
        for rfid in faculty.rfids.all(): 
            rfid.approved = 0
            rfid.save()
            #print(f"rfid : {rfid}, approve : {rfid.approved}")
           
    
    return -1


############################################################################
############################################################################
#API


@api_view(['POST'])
def check_rfid(request):
    RFID_input = request.data.get('RFID')
    local_tz = pytz.timezone('Asia/Manila')
    local_time = timezone.now().astimezone(local_tz)
    current_time = local_time.time()
    current_weekday = datetime.datetime.now().strftime("%A")

    weekday_map = {
    "Monday": "M",
    "Tuesday": "T",
    "Wednesday": "W",
    "Thursday": "R",
    "Friday": "F",
    "Saturday": "S",
    "Sunday": "U"
    }

    current_weekday_code = weekday_map.get(current_weekday)

    if not RFID_input:
           return Response({
            "status_message" : "Invalid Input"
        })

    try : 
        rfid = RFID.objects.get(rfid=RFID_input)

        response = {
            "status": "success",
            "message": "RFID exists",
            "ID": rfid.id,
            "RFID": rfid.rfid,
            "Approve": rfid.approved 
        }  
        
        if rfid.approved == 0 : 
            return Response(response)     
            
        scan = Scan.objects.filter(rfid=rfid).first()

        if not scan.faculty and not scan.student : 
            rfid.approved = 0 
            rfid.save()

            return Response(response)
        
        semester = Semester.objects.filter(isActive = True).first()

        if semester : 
            schedule = Schedule.objects.filter(
                start_time__lte=current_time,  
                end_time__gte=current_time,   
                weekdays__icontains=current_weekday_code, 
                semester = semester  
            ).first()

        if schedule : 
            print(10)
            class_object = ClassInstance.objects.filter(schedule = schedule, date = datetime2.today()).first()

            if not class_object : 
                class_object = ClassInstance(
                    schedule = schedule
                )
                class_object.save()
        
        if not schedule: 
            print(9)
            class_object = None

        if scan.faculty != None and scan.student == None :
            print(scan.faculty)
            user = scan.faculty
            type = "faculty"
            hasClass = False
            fullname = f"{scan.faculty.first_name} {scan.faculty.middle_initial}. {scan.faculty.last_name}"

            if schedule != None and schedule.faculty == scan.faculty :
                hasClass = True     

        elif scan.faculty == None and scan.student != None : 
            user = scan.student
            type = "student"
            hasClass = False
            fullname = f"{scan.student.first_name} {scan.student.middle_initial}. {scan.student.last_name}" 

            section = scan.student.section

            if schedule != None and schedule.section == section :
                hasClass = True

        logs = RfidLogs(
            user = fullname,
            type = type, 
            scan_time = datetime.datetime.now().strftime("%H:%M:%S")
        )

        logs.save()

        if class_object and schedule and hasClass == True: 
            attendance = Attendance.objects.filter(
                class_instance = class_object, 
                fullname = fullname, 
                type = type
            ).first()

            if not attendance :
                attendance = Attendance(
                    class_instance = class_object, 
                    fullname = fullname,
                    type = type, 
                    scan_time = datetime.datetime.now().strftime("%H:%M:%S")
                )

                attendance.save()

        if scan.faculty != None and scan.student == None : 
            computer = Computer.objects.filter(is_admin = 1).first()
            print(computer)
            normalize_macs = normalize_mac(computer.mac_address)
            send_magic_packet(normalize_macs)
            computer.status = 1
            computer.save()

        if scan.computer : 
            print(scan.computer.mac_address)
            normalize_macs = normalize_mac(scan.computer.mac_address)
            send_magic_packet(normalize_macs)
            scan.computer.status = 1
            scan.computer.save()

            
    except RFID.DoesNotExist:

        record = RFID.objects.create(rfid=RFID_input, approved=0)  # Default to 0 (denied)
        response = {
            "status": "success",
            "message": "RFID added successfully",
            "ID": record.id,
            "RFID": record.rfid,
            "Approve": record.approved
        }

    return Response(response)
 

# VERSION 2 
@api_view(['POST'])
# @permission_classes([IsAuthenticated])
def bind_rfid(request):  
    rfid = request.data.get('rfid')
    username = request.data.get('username', None)
    type = request.data.get('type', None)
    
    if not rfid:
        return Response({
          "status_message" : "Missing or invalid input"
        })
    
    rfid_object = RFID.objects.filter(rfid = rfid).first()

    if not rfid_object: 
        return Response({
            "status_message" : "Missing or invalid input"
        })
    
    scan = Scan.objects.filter(rfid = rfid_object).first()
    
    if scan: 
        scan.rfid = None
        scan.save()
        
    if (type != "student" and type != "faculty") or not username: 
        rfid_object.approved = 0 
        rfid_object.save()

        return Response({
            "status_message" : "Rfid is unassigned successfully"
        })
        
    if type == "student" :
        user = Student.objects.filter(username = username).first()

        if not user: 
            return Response({
            "status_message" : "Student does not exist"
            })
        
        scan = Scan.objects.filter(student__username = username, faculty__isnull = True).first()

        if not scan: 
            scan = Scan(
                student = user, 
                rfid = rfid_object
            )

        if scan : 
            scan.rfid = rfid_object

        scan.save()

        rfid_object.approved = 1
        rfid_object.save()

    if type == "faculty":
        user = User.objects.filter(username = username).first()

        if not user: 
            return Response({
            "status_message" : "Faculty does not exist"
            })
        
        scan = Scan.objects.filter(faculty__username = username).first()

        if not scan: 
            scan = Scan(
                faculty = user, 
                rfid = rfid_object
            )

        if scan : 
            scan.rfid = rfid_object

        scan.save()

        rfid_object.approved = 1
        rfid_object.save()

    return Response({
        "status_message" : "RFID was successfully binded to user"
    }) 
    
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_all_rfid(request):

    rfids = RFID.objects.all()
    
    response = []
    for rfid in rfids: 
        response.append(rfid.rfid)

    return Response({
        "status_message" : "RFID successfully retrieved",
        "rfids" : response
    }) 

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_logs_rfid(request): 
    pagination = request.data.get("pagination", None)
    start_date = request.data.get("start_date", None)
    end_date = request.data.get("end_date", None)
    username = request.data.get("username", None)
    type = request.data.get("type", None)

    if not start_date and not end_date :
        start_date_params = date(1, 1, 1)
        end_date_params = date(9999, 12, 31)

    elif not start_date:
        start_date_params = date(1, 1, 1)
        end_date_params = parse_date(end_date) 

    elif not end_date :
        start_date_params = parse_date(start_date)
        end_date_params = date(9999, 12, 31)

    else :
        start_date_params = parse_date(start_date)
        end_date_params = parse_date(end_date) 

    if not pagination: 
        pagination = 1
    
    pagination = float(pagination)
    
    logs = []
    
    items = int(pagination * 50) 
    
    filters = {
    'date__range': (start_date_params, end_date_params)
    }

    if username:
        filters['user__icontains'] = username

    if type : 
        filters['type'] = type

    logs = RfidLogs.objects.filter(**filters).order_by("-date", "-scan_time")[:items * 2]

    if pagination > 1 :
        excluded = items - 1 - 50
    else : 
        excluded = -1
    
    json_response = []
    for i in range(items):
        if i > excluded and i < len(logs):
            data = {
                "id" : logs[i].id,
                "username" : logs[i].user,
                "type" : logs[i].type,
                "date" : logs[i].date,
                "start_time" : logs[i].scan_time,
            }
            json_response.append(data)
    
    length = len(logs)
    pagination_length = length / 50

    if pagination_length % 1 != 0:
        pagination_length = math.floor(pagination_length)
        pagination_length += 1  

           
    return Response({
        "status_message" : "Logs obtained succesfully",
        "logs" : json_response,
        "pagination_length" : pagination_length
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def delete_rfid(request): 
    rfid = request.data.get("rfid")

    if not rfid:
        return Response({
            "status_message" : "Rfid is required"
        })    
    
    rfid = RFID.objects.filter(rfid=rfid)

    if not rfid: 
        return Response({
            "status_message" : "RFID not found"
        })
    
    rfid.delete()

    return Response({
        "status_message" : "RFID deleted successfully"
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_attendance_info(request): 
    schedule_id = request.data.get("schedule_id")
    sort_by = request.data.get("sort_by", None)

    if not schedule_id and not isinstance(schedule_id, int):
        return Response({
            "status_message" : "Invalid or missing input"
        })
    
    
    schedule = Schedule.objects.filter(id = schedule_id).first()

    if not schedule : 
        return Response({
            "status_message" : "Schedule does not exist"
        })
    
    
    attendance_response = []
    class_response = []

    section = schedule.section

    classes = ClassInstance.objects.filter(schedule = schedule)

    for class_object in classes : 
        class_response.append(class_object.date)
        attendances = Attendance.objects.filter(class_instance = class_object)
        students = Student.objects.filter(section = section)
        faculty_attendance = Attendance.objects.filter(class_instance__schedule__faculty = schedule.faculty).first()
        faculty = schedule.faculty
        fullname = f"{faculty.first_name} {faculty.middle_initial}. {faculty.last_name}"
        if faculty_attendance: 
            faculty_response = {
                "fullname" : fullname, 
                "log_time" : faculty_attendance.scan_time, 
                "type" : "faculty"
                }
            
        else : 
            faculty_response = {
                "fullname" : fullname, 
                "log_time" : "Did not attend", 
                "type" : "faculty"
            }
            
        attendees = []

        for student in students : 
            data = None
            for attendance in attendances:
                if attendance.fullname == f"{student.first_name} {student.middle_initial}. {student.last_name}":
                    formatted_start_time = datetime2.combine(datetime2.today(), schedule.start_time)
                    formatted_scan_time = datetime2.combine(datetime2.today(), attendance.scan_time)
                    if formatted_scan_time > formatted_start_time + timedelta(minutes=15):
                        attendance_position = "late"
                    
                    elif formatted_scan_time > formatted_start_time + timedelta(minutes=30):
                        attendance_position = "absent"

                    else : 
                        attendance_position = "present"

                    name = format_fullname_lastname_first(attendance.fullname)

                    data = {
                        "fullname": name,
                        "log_time": attendance.scan_time,
                        "type": attendance.type, 
                        "attendance_position" : attendance_position
                    }
                    attendees.append(data)
            
            if data == None : 
                data = {
                    "fullname" : f"{student.last_name}, {student.first_name} {student.middle_initial}. ",
                    "log_time": None,
                    "type": "student", 
                    "attendance_position" : "absent"
                }
                attendees.append(data)

        if sort_by == "asc_by_name" : 
            attendees.sort(key=lambda x: x["fullname"].lower())

        elif sort_by == "desc_by_name" : 
            attendees.sort(reverse=True, key=lambda x: x["fullname"].lower())

        elif sort_by == "asc_by_time" or sort_by == None: 
            attendees.sort(key=lambda x: (
                datetime.strptime(x["log_time"], '%H:%M:%S').time() if isinstance(x["log_time"], str) 
                else (x["log_time"] if x["log_time"] is not None else datetime2.max.time()),  
                x["fullname"]
            ))
        
        elif sort_by == "desc_by_time" :
            attendees.sort(reverse=True, key=lambda x: (
                datetime.strptime(x["log_time"], '%H:%M:%S').time() if isinstance(x["log_time"], str) 
                else (x["log_time"] if x["log_time"] is not None else datetime2.max.time()),  
                x["fullname"]
            ))

        
        
        attendance_data = {
            "date" : class_object.date, 
            "date_id" : class_object.id,
            "faculty" : faculty_response, 
            "attendees" : attendees
        }

        attendance_response.append(attendance_data)

    print(attendance_response)
    return Response({
        "date" : class_response,
        "attendance" : attendance_response
    })


@api_view(['POST'])
# @permission_classes([IsAuthenticated])
def get_cumulative_attendance(request): 
    schedule_id = request.data.get("schedule_id")

    if not schedule_id : 
        return Response({
            "status" : 0, 
            "status_message" : "Missing or invalid input"
        })

    schedule = Schedule.objects.filter(id = schedule_id).first()

    if not schedule: 
        return Response({
            "status" : 0, 
            "status_message" : "Schedule does not exist"
        })

    section = schedule.section

    if not section : 
        return Response({
            "status" : 0, 
            "status_message" : "Section does not exist"
        })

    students = Student.objects.filter(section = section).order_by("last_name", "first_name", "middle_initial")

    response = []
    students_response = []
    class_instance_response = []

    for student in students: 
        fullname = format_fullname(student.first_name, student.middle_initial, student.last_name)
        class_instances = ClassInstance.objects.filter(schedule = schedule)

        present_classes = [] 
        number_present = 0 
        absent_classes = []
        number_absent = 0 
        late_classes = []
        number_late = 0

        for instance in class_instances: 
            attendance = Attendance.objects.filter(fullname = fullname, class_instance = instance).first()

            if not attendance : 
                absent_classes.append(instance.date)
                number_absent += 1
                continue
            
            formatted_start_time = datetime2.combine(datetime2.today(), schedule.start_time)
            formatted_scan_time = datetime2.combine(datetime2.today(), attendance.scan_time)
            
            if formatted_scan_time > formatted_start_time + timedelta(minutes=15):
                late_classes.append(instance.date)
                number_late += 1
                continue

            elif formatted_scan_time > formatted_start_time + timedelta(minutes=30):
                attendance_position = "absent"
                number_absent += 1
                continue
            
            else : 
                present_classes.append(instance.date)
                number_present += 1
                continue
        
        data = {
            "fullname" : format_fullname_lastname_first(fullname), 
            "absents" : absent_classes,
            "number_absents" : number_absent,
            "lates" : late_classes,
            "number_lates" : number_late, 
            "presents" : present_classes,
            "number_presents" : number_present
        }

        students_response.append(data)

    number_classes = 0
    for instance in class_instances:
        class_instance_response.append(instance.date)
        number_classes += 1

    students_response.sort(key=lambda x: x["fullname"])

    data = {
        "status" : 1, 
        "status_message" : "Cumulative attendance has been obtained successfully", 
        "dates" : class_instance_response, 
        "number_classes" : number_classes, 
        "attendees" : students_response
    }

    return Response(data)


def apply_student_name_and_attendance_status_style(sheet, start_row=2, max_row=None):
    # Apply a consistent style to student names and attendance statuses
    light_blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Very Light Blue
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")      # White
    
    present_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")  # Green for Present
    late_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")     # Yellow for Late
    absent_fill = PatternFill(start_color="FF8A80", end_color="FF8A80", fill_type="solid")   # Red for Absent

    for row in sheet.iter_rows(min_row=start_row, max_row=max_row):
        for cell in row:
            if cell.column == 1:  # Apply styling to the "Student Name" column
                # Alternate row colors for Student Names
                if cell.row % 2 == 0:  # Even rows
                    cell.fill = light_blue_fill
                else:  # Odd rows
                    cell.fill = white_fill
                # Align student names to the left
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # Set font color
                cell.font = Font(color="000000")
                # Add borders
                cell.border = Border(bottom=Side(border_style="thin", color="D3D3D3"))

            # Apply colors for attendance status (Present, Late, Absent)
            elif cell.column > 1:  # If it's a cell in the attendance status columns
                if cell.value == "Present":
                    cell.fill = present_fill
                elif cell.value == "Late":
                    cell.fill = late_fill
                elif cell.value == "Absent":
                    cell.fill = absent_fill
                # Align to center and set borders
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=Side(border_style="thin", color="D3D3D3"))

@api_view(['GET'])
def generate_cumulative_and_summary_reports(request):
    schedule_id = request.query_params.get("schedule_id")

    if not schedule_id:
        return Response({"status_message": "Missing or invalid schedule_id"}, status=400)

    schedule = Schedule.objects.filter(id=schedule_id).first()
    if not schedule:
        return Response({"status_message": "Schedule does not exist"}, status=400)

    section = schedule.section
    students = Student.objects.filter(section=section).order_by("last_name", "first_name", "middle_initial")
    class_instances = ClassInstance.objects.filter(schedule=schedule)

    total_classes = class_instances.count()
    faculty_name = f"{schedule.faculty.first_name} {schedule.faculty.middle_initial}. {schedule.faculty.last_name}"
    subject_name = schedule.subject

    cumulative_data = []
    summary_data = []

    for student in students:
        fullname = format_fullname(student.first_name, student.middle_initial, student.last_name)
        attendance_statuses = []
        days_present, days_late, days_absent = 0, 0, 0

        for class_instance in class_instances:
            attendance = Attendance.objects.filter(fullname=fullname, class_instance=class_instance).first()

            if not attendance:
                attendance_statuses.append("Absent")
                days_absent += 1
                continue

            formatted_start_time = datetime2.combine(datetime2.today(), schedule.start_time)
            formatted_scan_time = datetime2.combine(datetime2.today(), attendance.scan_time)

            if formatted_scan_time > formatted_start_time + timedelta(minutes=15):
                attendance_statuses.append("Late")
                days_late += 1
            elif formatted_scan_time <= formatted_start_time + timedelta(minutes=15):
                attendance_statuses.append("Present")
                days_present += 1
            elif formatted_scan_time > formatted_start_time + timedelta(minutes=30):
                attendance_statuses.append("Absent")
                days_absent += 1

        cumulative_data.append([format_fullname_lastname_first(fullname)] + attendance_statuses)
        summary_data.append([format_fullname_lastname_first(fullname), days_present, days_late, days_absent])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_reports.xlsx'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        # Cumulative Attendance Report
        cumulative_df = pd.DataFrame(
            cumulative_data,
            columns=["Student Name"] + [ci.date.strftime("%Y-%m-%d") for ci in class_instances]
        )
        cumulative_df.to_excel(writer, index=False, sheet_name="Cumulative Attendance")
        cumulative_sheet = writer.sheets["Cumulative Attendance"]
        cumulative_sheet.insert_rows(0)
        cumulative_sheet["A1"] = f"Subject: {subject_name}    Total Classes: {total_classes}    Faculty: {faculty_name}"

        # Apply header style
        for cell in cumulative_sheet[1]:
            cell.fill = PatternFill(start_color="3A6B8D", end_color="3A6B8D", fill_type="solid")  # Soft Blue
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color="FFFFFF")  # Ensure font is white
            cell.border = Border(bottom=Side(border_style="thin", color="000000"))

        # Apply bold and left-aligned style for "Student Name" header (A1)
        student_name_cell = cumulative_sheet['A1']
        student_name_cell.font = Font(bold=True, color="FFFFFF")  # Set font to white
        student_name_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Apply student name and attendance status styling using the new helper function
        apply_student_name_and_attendance_status_style(cumulative_sheet)

        # Automatically adjust column widths
        for col in range(1, len(cumulative_data[0]) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in cumulative_sheet.iter_rows(min_row=1, max_row=cumulative_sheet.max_row, min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)  # Add some padding to the column width
            cumulative_sheet.column_dimensions[column].width = adjusted_width

        # Summary Attendance Report
        summary_df = pd.DataFrame(
            summary_data,
            columns=["Student Name", "Days Present", "Days Late", "Days Absent"]
        )
        summary_df.to_excel(writer, index=False, sheet_name="Summary Attendance")
        summary_sheet = writer.sheets["Summary Attendance"]
        summary_sheet.insert_rows(0)
        summary_sheet["A1"] = f"Subject: {subject_name}    Total Classes: {total_classes}    Faculty: {faculty_name}"

        # Apply header style for summary sheet
        for cell in summary_sheet[1]:
            cell.fill = PatternFill(start_color="3A6B8D", end_color="3A6B8D", fill_type="solid")  # Soft Blue
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color="FFFFFF")  # Ensure font is white
            cell.border = Border(bottom=Side(border_style="thin", color="000000"))

        # Apply bold and left-aligned style for "Student Name" header (A1) in summary sheet
        student_name_cell_summary = summary_sheet['A1']
        student_name_cell_summary.font = Font(bold=True, color="FFFFFF")  # Set font to white
        student_name_cell_summary.alignment = Alignment(horizontal="left", vertical="center")

        # Apply student name and attendance status styling using the new helper function
        apply_student_name_and_attendance_status_style(summary_sheet)

        # Automatically adjust column widths for the summary sheet
        for col in range(1, len(summary_data[0]) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row, min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)  # Add some padding to the column width
            summary_sheet.column_dimensions[column].width = adjusted_width

    return response


@api_view(['POST'])
# @permission_classes([IsAuthenticated])            
def get_summary_attendance(request):
    schedule_id = request.data.get("schedule_id")

    if not schedule_id : 
        return Response({
            "status" : 0, 
            "status_message" : "Missing or invalid input"
        })

    schedule = Schedule.objects.filter(id = schedule_id).first()

    if not schedule: 
        return Response({
            "status" : 0, 
            "status_message" : "Schedule does not exist"
        })

    section = schedule.section

    if not section : 
        return Response({
            "status" : 0, 
            "status_message" : "Section does not exist"
        })

    students = Student.objects.filter(section = section).order_by("last_name", "first_name", "middle_initial")

    response = []
    students_response = []
    number_of_class_instance = 0

    for student in students: 
        fullname = format_fullname(student.first_name, student.middle_initial, student.last_name)
        class_instances = ClassInstance.objects.filter(schedule = schedule)

        present_classes = 0
        absent_classes = 0
        late_classes = 0

        for instance in class_instances: 
            attendance = Attendance.objects.filter(fullname = fullname, class_instance = instance).first()

            if not attendance : 
                absent_classes += 1
                continue
            
            formatted_start_time = datetime2.combine(datetime2.today(), schedule.start_time)
            formatted_scan_time = datetime2.combine(datetime2.today(), attendance.scan_time)
            
            if formatted_scan_time > formatted_start_time + timedelta(minutes=15):
                late_classes += 1
                continue
            
            else : 
                present_classes += 1

        data = {
            "fullname" : fullname, 
            "absents" : absent_classes,
            "lates" : late_classes,
            "presents" : present_classes
        }

        students_response.append(data)

    for instance in class_instances:
        number_of_class_instance += 1

    data = {
        "status" : 1, 
        "status_message" : "Cumulative attendance has been obtained successfully", 
        "number_of_class" : number_of_class_instance, 
        "attendees" : students_response
    }

    return Response(data)

            



    






    
