import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from django.http import HttpResponse
from datetime import date
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from .models import UserLog, RfidLogs
from .models import Computer, Semester, Schedule, RfidLogs, ClassInstance, Attendance, User, RFID, User, Scan, Student
from .forms import ScheduleForm
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from datetime import datetime as datetime2
from django.utils.dateparse import parse_date
import math
import datetime
from datetime import timedelta
from .utils import format_fullname, format_fullname_lastname_first


def parse_date(date_str):
    """ Parse date string in 'YYYY-MM-DD' format and return a date object. """
    return timezone.datetime.strptime(date_str, '%Y-%m-%d').date()

def get_default_dates():
    """ Return default start and end dates if not provided in the query. """
    today = date.today()
    start_date = date(today.year, today.month, 1)  # Default to the 1st of the current month
    end_date = today  # Default to today's date
    return start_date, end_date

def set_column_widths(sheet, dataframe):
    """ Automatically adjust column widths based on the max length of the data in each column. """
    for idx, col in enumerate(dataframe.columns, 1):
        max_length = max(dataframe[col].astype(str).apply(len).max(), len(col))
        sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max_length + 2  # Add some padding

def apply_borders(sheet):
    """ Apply borders to all cells in the sheet. """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

def format_headers(sheet):
    """ Format header row with bold font and centered alignment. """
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

def format_dates(sheet, date_columns):
    """ Format date columns in the sheet. """
    for col in date_columns:
        for cell in sheet[col]:
            if isinstance(cell.value, date):
                cell.number_format = 'YYYY-MM-DD'  # Format as date
            elif isinstance(cell.value, str) and 'AM' in cell.value or 'PM' in cell.value:
                cell.alignment = Alignment(horizontal='center')

# Faculty Only Report
@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def generate_faculty_report_excel(request):
    # Get query parameters for start_date, end_date
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')

    # If no date is provided, use default dates
    if not start_date_str or not end_date_str:
        start_date, end_date = get_default_dates()
    else:
        # Parse the start and end dates
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)

    # Filter UserLogs for faculty only
    user_logs = UserLog.objects.filter(
        section="faculty",  # Only include faculty logs
        date__range=[start_date, end_date]  # Filter by date range
    ).order_by('-date', '-logonTime')  # Ordering by date and logonTime in descending order

    # Prepare data for the report
    data = []
    for log in user_logs:
        row = [
            log.computer if log.computer else 'N/A',  # Computer used
            log.user,  # Full name
            log.section,  # Section (Faculty)
            log.date,  # Date of usage
            log.logonTime,  # Start of usage
            log.logoffTime if log.logoffTime else 'Not yet Logged off',  # End of usage
        ]
        data.append(row)

    # Convert data to DataFrame
    df = pd.DataFrame(data, columns=["Computer", "User", "Section", "Date of Usage", "Login", "Logout"])

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=faculty_report_{start_date}_{end_date}.xlsx'

    # Write the DataFrame to the Excel file
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Faculty Log Report')

        # Get the sheet
        sheet = writer.sheets['Faculty Log Report']
        
        # Set column widths
        set_column_widths(sheet, df)

        # Apply formatting to headers
        format_headers(sheet)

        # Apply borders to all cells
        apply_borders(sheet)

        # Format date columns (Logon Date and Logoff Time)
        format_dates(sheet, date_columns=['D', 'F'])

    return response

# Student Only Report
@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def generate_student_report_excel(request):
    # Get query parameters for start_date, end_date, and section
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')
    section = request.query_params.get('section', '')  # Default to empty string if not specified

    # If no date is provided, use default dates
    if not start_date_str or not end_date_str:
        start_date, end_date = get_default_dates()
    else:
        # Parse the start and end dates
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)

    # Base filter for UserLogs by date range (exclude faculty logs)
    filters = {
        'date__range': [start_date, end_date],  # Filter by date range
    }

    # Exclude logs where section is 'faculty'
    user_logs = UserLog.objects.exclude(
        section="faculty"  # Exclude faculty logs
    ).filter(**filters)  # Apply the other filters (like date range)

    # If section is specified, filter by that section; if not, include all sections
    if section:
        user_logs = user_logs.filter(section=section)  # Filter by section name (e.g., 'A', 'B')

    # Ordering by date and logonTime
    user_logs = user_logs.order_by('-date', '-logonTime')

    # Prepare data for the report
    data = []
    for log in user_logs:
        row = [
            log.computer if log.computer else 'N/A',  # Computer used
            log.user,  # Full name
            log.section,  # Section (Faculty)
            log.date,  # Date of usage
            log.logonTime,  # Start of usage
            log.logoffTime if log.logoffTime else 'Not yet Logged off',  # End of usage
        ]
        data.append(row)

    # Convert data to DataFrame
    df = pd.DataFrame(data, columns=["Computer", "User", "Section", "Date of Usage", "Login", "Logout"])

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=student_report_{start_date}_{end_date}.xlsx'

    # Write the DataFrame to the Excel file
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Student Log Report')

        # Get the sheet
        sheet = writer.sheets['Student Log Report']
        
        # Set column widths
        set_column_widths(sheet, df)

        # Apply formatting to headers
        format_headers(sheet)

        # Apply borders to all cells
        apply_borders(sheet)

        # Format date columns (Logon Date and Logoff Time)
        format_dates(sheet, date_columns=['D', 'F'])

    return response

# Combined Report (Faculty + Student)
@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def generate_combined_report_excel(request):
    # Get query parameters for start_date, end_date
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')

    # If no date is provided, use default dates
    if not start_date_str or not end_date_str:
        start_date, end_date = get_default_dates()
    else:
        # Parse the start and end dates
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)

    # Filter UserLogs for both faculty and student within the date range
    user_logs = UserLog.objects.filter(
        date__range=[start_date, end_date]  # Filter by date range
    ).order_by('-date', '-logonTime')  # Ordering by date and logonTime in descending order

    # Prepare data for the combined report
    data = []
    for log in user_logs:
        row = [
            log.computer if log.computer else 'N/A',  # Computer used
            log.user,  # Full name
            log.section,  # Section (Faculty)
            log.date,  # Date of usage
            log.logonTime,  # Start of usage
            log.logoffTime if log.logoffTime else 'Not yet Logged off',  # End of usage
        ]
        data.append(row)

    # Convert data to DataFrame
    df = pd.DataFrame(data, columns=["Computer", "User", "Section", "Date of Usage", "Login", "Logout"])

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=combined_report_{start_date}_{end_date}.xlsx'

    # Write the DataFrame to the Excel file
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Combined Log Report')

        # Get the sheet
        sheet = writer.sheets['Combined Log Report']
        
        # Set column widths
        set_column_widths(sheet, df)

        # Apply formatting to headers
        format_headers(sheet)

        # Apply borders to all cells
        apply_borders(sheet)

        # Format date columns (Logon Date and Logoff Time)
        format_dates(sheet, date_columns=['D', 'F'])

    return response


@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def generate_rfid_report_excel(request):
    # Get query parameters for start_date, end_date, and type
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')
    rfid_type = request.query_params.get('type', '')  # Default to empty string if not specified

    # If no date is provided, use default dates
    if not start_date_str or not end_date_str:
        start_date, end_date = get_default_dates()
    else:
        # Parse the start and end dates
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)

    # Base filter for RfidLogs by date range
    filters = {
        'date__range': [start_date, end_date],  # Filter by date range
    }

    # If type is specified (either 'student' or 'faculty'), filter by that type
    if rfid_type:
        filters['type'] = rfid_type  # Filter by type (either 'student' or 'faculty')

    # Query RfidLogs with the dynamic filters
    rfid_logs = RfidLogs.objects.filter(**filters).order_by('-date', '-scan_time')

    # Prepare data for the report
    data = []
    for log in rfid_logs:
        row = [
            log.user,  # User (full name or identifier)
            log.type.capitalize(),  # Type ('student' or 'faculty')
            log.date,  # Date of scan
            log.scan_time  # Scan time
        ]
        data.append(row)

    # Convert data to DataFrame
    df = pd.DataFrame(data, columns=["User", "Type", "Scan Date", "Scan Time"])

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rfid_report_{start_date}_{end_date}.xlsx'

    # Write the DataFrame to the Excel file
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='RFID Log Report')

        # Get the sheet
        sheet = writer.sheets['RFID Log Report']
        
        # Set column widths
        set_column_widths(sheet, df)

        # Apply formatting to headers
        format_headers(sheet)

        # Apply borders to all cells
        apply_borders(sheet)

        # Format date columns (Scan Date)
        format_dates(sheet, date_columns=['C'])

    return response

def parse_log_time(log_time):
    """Convert log_time to a datetime object, handling 'Did not attend' or None."""
    if log_time == "Did not attend" or log_time is None:
        return None  # or return a default time if needed (e.g., datetime.time(23, 59))
    if isinstance(log_time, str):
        try:
            # If log_time is a string, convert it to a time object (assuming format HH:MM:SS)
            return datetime.strptime(log_time, "%H:%M:%S").time()
        except ValueError:
            return None  # Handle invalid time format gracefully
    return log_time  # If it's already a datetime.time object, return it as-is

def sort_dataframe(df, sort_by):
    """Sort the dataframe by the given sort type."""
    # Convert log_time to datetime objects for comparison
    df['Log Time'] = df['Log Time'].apply(parse_log_time)

    if sort_by == "asc_by_name":
        df = df.sort_values(by="Name", ascending=True)
    elif sort_by == "desc_by_name":
        df = df.sort_values(by="Name", ascending=False)
    elif sort_by == "asc_by_time":
        df = df.sort_values(by="Log Time", ascending=True, na_position='last')  # Ensure NaT values are handled
    elif sort_by == "desc_by_time":
        df = df.sort_values(by="Log Time", ascending=False, na_position='last')  # Ensure NaT values are handled
    return df

@api_view(['GET'])
def generate_attendance_report_excel(request):
    # Get filters from request
    schedule_id = request.query_params.get("schedule_id")
    semester_id = request.query_params.get("semester_id")
    schedule_date = request.query_params.get("schedule_date")
    sort_by = request.query_params.get("sort_by", "asc_by_name")  # Default sort is ascending by name

    # Initialize queryset for class instances
    class_instances = ClassInstance.objects.all()

    # Apply filters (same as before)
    if semester_id:
        semester = Semester.objects.filter(id=semester_id).first()
        if not semester:
            return Response({"status_message": "Semester does not exist"}, status=400)
        schedules_in_semester = Schedule.objects.filter(semester=semester)
        class_instances = class_instances.filter(schedule__in=schedules_in_semester)

    if schedule_id:
        schedule = Schedule.objects.filter(id=schedule_id).first()
        if not schedule:
            return Response({"status_message": "Schedule does not exist"}, status=400)
        class_instances = class_instances.filter(schedule=schedule)

    if schedule_date:
        class_instances = class_instances.filter(date=schedule_date)

    if not class_instances:
        return Response({"status_message": "No class instances found for the given filters"}, status=404)

    # Create HTTP response to serve the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_report.xlsx'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        all_attendance_data = []  # List to hold all the data for each class instance

        for class_instance in class_instances:
            date = class_instance.date
            schedule = class_instance.schedule
            section = schedule.section
            subject = schedule.subject

            # Prepare data for the current class instance
            attendance_data = []

            faculty_attendance = Attendance.objects.filter(class_instance=class_instance, type="faculty").first()
            if faculty_attendance:
                faculty_fullname = f"{schedule.faculty.last_name}, {schedule.faculty.first_name} {schedule.faculty.middle_initial}."
                faculty_entry = {
                    "fullname": faculty_fullname,
                    "log_time": faculty_attendance.scan_time,
                    "type": "faculty"
                }
            else:
                faculty_entry = {
                    "fullname": f"{schedule.faculty.last_name}, {schedule.faculty.first_name} {schedule.faculty.middle_initial}.",
                    "log_time": "Did not attend",
                    "type": "faculty"
                }

            students = Student.objects.filter(section=section)
            student_attendees = []
            for student in students:
                attendance = Attendance.objects.filter(class_instance=class_instance, fullname=f"{student.first_name} {student.middle_initial}. {student.last_name}").first()
                if attendance:
                    student_entry = {
                        "fullname": format_fullname_lastname_first(attendance.fullname),
                        "log_time": attendance.scan_time,
                        "type": attendance.type
                    }
                else:
                    student_entry = {
                        "fullname": f"{student.last_name}, {student.first_name} {student.middle_initial}.",
                        "log_time": "Did not attend",
                        "type": "student"
                    }
                student_attendees.append(student_entry)

            # Combine the attendance data
            attendance_data.append({
                "date": date,
                "subject": subject,
                "faculty": faculty_entry,
                "attendees": student_attendees
            })

            # Prepare the data for DataFrame
            data = []
            for record in attendance_data:
                for attendee in record['attendees']:
                    data.append([record['date'], record['subject'], attendee['fullname'], attendee['log_time'], attendee['type']])

            # Convert to DataFrame
            df = pd.DataFrame(data, columns=["Date", "Subject", "Name", "Log Time", "Type"])

            # Sort data based on the 'sort_by' parameter
            df = sort_dataframe(df, sort_by)

            # Use a unique sheet name based on the class instance's date and subject
            sheet_name = f"{class_instance.date} - {class_instance.schedule.subject} - {class_instance.id}"

            # Write DataFrame to Excel sheet
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            sheet = writer.sheets[sheet_name]

            # Set column widths and formatting (same as before)
            set_column_widths(sheet, df)
            format_headers(sheet)
            apply_borders(sheet)

    return response