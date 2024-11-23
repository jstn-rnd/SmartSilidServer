from django.shortcuts import render
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
from datetime import datetime
from .models import Scan, User, Student, RfidLogs

# Function to automatically adjust column widths based on the content
def set_column_widths(sheet, data):
    """ Automatically adjust column widths based on the max length of the data in each column. """
    for col_idx, column in enumerate(data, 1):
        max_length = max(len(str(item)) for item in column)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2

# Function to apply borders to all cells
def apply_borders(sheet):
    """ Apply borders to all cells in the sheet. """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

# Function to format header with bold and centered alignment
def format_headers(sheet):
    """ Format header row with bold font and centered alignment. """
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Function to format date columns in the sheet
def format_dates(sheet, date_columns):
    """ Format date columns in the sheet. """
    for col in date_columns:
        for cell in sheet[col]:
            if isinstance(cell.value, datetime):
                cell.number_format = 'YYYY-MM-DD'  # Format as date

# Generate Faculty RFID Report
def generate_faculty_rfid_report(request):
    start_date = request.GET.get('start_date', None)
    end_date = request.GET.get('end_date', None)
    
    scans = Scan.objects.filter(faculty__type='faculty')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        scans = scans.filter(date__gte=start_date)
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        scans = scans.filter(date__lte=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty RFID Report"

    # Define header for the report
    header = ["Full Name", "Type", "Scan Date", "Scan Time"]
    ws.append(header)

    data = []

    for scan in scans:
        faculty = scan.faculty
        scan_date = None
        scan_time = None

        rfid_logs = RfidLogs.objects.filter(user=faculty.username, type="faculty")
        
        if rfid_logs.exists():
            rfid_log = rfid_logs.first()
            scan_date = rfid_log.date
            scan_time = rfid_log.scan_time

        data.append([
            f"{faculty.first_name} {faculty.middle_initial}. {faculty.last_name}", 
            "Faculty",
            scan_date,
            scan_time
        ])

    for row in data:
        ws.append(row)

    # Apply formatting to the sheet
    set_column_widths(ws, zip(*data))  # Adjust column widths
    format_headers(ws)  # Bold and center headers
    apply_borders(ws)  # Apply borders to all cells
    format_dates(ws, date_columns=['C', 'D'])  # Format date columns

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=faculty_rfid_report.xlsx'
    wb.save(response)
    return response

# Generate Student RFID Report
def generate_student_rfid_report(request):
    start_date = request.GET.get('start_date', None)
    end_date = request.GET.get('end_date', None)
    
    scans = Scan.objects.filter(student__isnull=False)
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        scans = scans.filter(date__gte=start_date)
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        scans = scans.filter(date__lte=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Student RFID Report"

    header = ["Full Name", "Type", "Scan Date", "Scan Time"]
    ws.append(header)

    data = []

    for scan in scans:
        student = scan.student
        scan_date = None
        scan_time = None

        rfid_logs = RfidLogs.objects.filter(user=student.username, type="student")
        
        if rfid_logs.exists():
            rfid_log = rfid_logs.first()
            scan_date = rfid_log.date
            scan_time = rfid_log.scan_time

        data.append([
            f"{student.first_name} {student.middle_initial}. {student.last_name}",
            "Student",
            scan_date,
            scan_time
        ])

    for row in data:
        ws.append(row)

    # Apply formatting to the sheet
    set_column_widths(ws, zip(*data))  # Adjust column widths
    format_headers(ws)  # Bold and center headers
    apply_borders(ws)  # Apply borders to all cells
    format_dates(ws, date_columns=['C', 'D'])  # Format date columns

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=student_rfid_report.xlsx'
    wb.save(response)
    return response

# Generate Combined RFID Report
def generate_combined_rfid_report(request):
    start_date = request.GET.get('start_date', None)
    end_date = request.GET.get('end_date', None)
    
    scans = Scan.objects.all()
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        scans = scans.filter(date__gte=start_date)
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        scans = scans.filter(date__lte=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Combined RFID Report"

    header = ["Full Name", "Type", "Scan Date", "Scan Time"]
    ws.append(header)

    data = []

    for scan in scans:
        if scan.faculty:
            name = f"{scan.faculty.first_name} {scan.faculty.middle_initial}. {scan.faculty.last_name}"
            type = "Faculty"
            rfid_logs = RfidLogs.objects.filter(user=scan.faculty.username, type="faculty")
        elif scan.student:
            name = f"{scan.student.first_name} {scan.student.middle_initial}. {scan.student.last_name}"
            type = "Student"
            rfid_logs = RfidLogs.objects.filter(user=scan.student.username, type="student")
        else:
            continue
        
        scan_date = None
        scan_time = None

        if rfid_logs.exists():
            rfid_log = rfid_logs.first()
            scan_date = rfid_log.date
            scan_time = rfid_log.scan_time
        
        data.append([name, type, scan_date, scan_time])

    for row in data:
        ws.append(row)

    # Apply formatting to the sheet
    set_column_widths(ws, zip(*data))  # Adjust column widths
    format_headers(ws)  # Bold and center headers
    apply_borders(ws)  # Apply borders to all cells
    format_dates(ws, date_columns=['C', 'D'])  # Format date columns

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=combined_rfid_report.xlsx'
    wb.save(response)
    return response